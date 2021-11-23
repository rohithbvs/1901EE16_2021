import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import re

def feedback_not_submitted():

	ltp_mapping_feedback_type = {1: 'lecture', 2: 'tutorial', 3:'practical'}
	output_file_name = "course_feedback_remaining.xlsx"

	'''name_roll is dictionary which has roll no as key and email,aemail,contact,name coressponding to that rollno in a list ''' 
	''' {'1811MS10': ['chris_trent834@gembat.biz', 'chris_trent6297@twace.org', '9876543210', 'Chris Trent'] this is an example how it stores'''
	name_roll = {}
	with open('studentinfo.csv','r') as file:
		reader = csv.DictReader(file)
		for row in reader:
			name_roll[row['Roll No']] = [row['email'],row['aemail'],row['contact'],row['Name']]
	

	'''roll_course_enrolled is a dictionary which has rollno as a key and stores a 2d list of all subcodes he register'''
	''' '1901CE05': [['5', '5', 'CE303'], ['5', '5', 'CE317'], ['5', '5', 'CE319'], ['5', '5', 'CE321'], ['5', '5', 'CE391'], 
	['5', '5', 'CE393'], ['5', '5', 'HS301'], ['5', '3', 'CE291'], ['5', '3', 'CE293'], ['5', '2', 'EE103'], ['5', '4', 'CE292']]'''
	'''the above line is example how row_course_enrolled is storing roll no corresponding students courses enrolled'''

	roll_course_enrolled = {}
	with open('course_registered_by_all_students.csv','r') as file:
		reader = csv.DictReader(file)
		for row in reader:
			if row['rollno'] in roll_course_enrolled:
				roll_course_enrolled[row['rollno']].append([row['register_sem'],row['schedule_sem'],row['subno']])
			else:
				roll_course_enrolled[row['rollno']] = []
				roll_course_enrolled[row['rollno']].append([row['register_sem'],row['schedule_sem'],row['subno']])


	'''feedback_submitted is a dictionary rollno as key and 2dlist which stores all feedback submitted by students for different subcode'''
	''' 1801ME51': [['Abdul Windsor', 'ME495', '3'], ['Abdul Windsor', 'ME501', '1'], ['Abdul Windsor', 'ME504', '1'], ['Abdul Windsor', 'PH403', '1'], ['Abdul Windsor', 'ME431', '1']]'''
	''' the above line is example how feedback_submitted stores the data '''

	feedback_submitted = {}
	with open('course_feedback_submitted_by_students.csv','r') as file:
		reader = csv.DictReader(file)
		for row in reader:
			if row['stud_roll'] in feedback_submitted:
				#print(row['stud_roll'])
				feedback_submitted[row['stud_roll']].append([row['stud_name'],row['course_code'],row['feedback_type']])
			else:
				#print(row['stud_roll'])
				feedback_submitted[row['stud_roll']] = []
				feedback_submitted[row['stud_roll']].append([row['stud_name'],row['course_code'],row['feedback_type']])
				#print(feedback_submitted[])

	'''subcode_ltp is a dictionary with subcode as key and list of set bits in ltp'''
	''''CB201': [1, 2] subcode_ltp stores in the following way stating sub CB102 has lectures and tutorial and doesn't have practical'''
	'''i have regrex groups for extracting l,t,p values seperately and checked whether they are 0 or not'''

	subcode_ltp={}
	with open('course_master_dont_open_in_excel.csv','r') as file:
		reader = csv.DictReader(file)
		for row in reader:
			lst = []
			regex = r"([0-9.]+)-([0-9.]+)-([0-9.]+)"
			match = re.search(regex, row['ltp'])
			if match != None:
				if str(match.group(1)) != '0':
					lst.append(1)
					
				if str(match.group(2)) != '0':
					lst.append(2)
					
				if str(match.group(3)) != '0':
					lst.append(3)
			
			subcode_ltp[row['subno']] = lst 

	#print(name_roll)
	#print(roll_course_enrolled)
	#print(feedback_submitted)
	#print(subcode_ltp)

	'''ans is the final list which contains the students roll and the subcode of missed feedback'''
	'''i have compared the students across the course_registered_by_all_students roll no to the feedback_submitted '''
	'''first i took a roll no from registered course csv(roll_course_enrolled) and then i took a subcode(subcode=row[2]) of that student enrolled'''
	'''then i took the ltp corresponding to the subcode(subcode_ltp[subcode]) and checked it with the feedback data'''
	'''in feedback data first i checked if roll no exist or not then i checked if the subcode that we initially took for registered csv is present in feedback or not'''
	'''finally i checked feedback type with ltp'''
	'''if there exist a subcode and feedbacktype then flag is true else flag is false and loop is breaked and we append in ans list'''
	ans = []
	for roll in roll_course_enrolled:
		for row in roll_course_enrolled[roll]:
			subcode = row[2]
			lst2 = subcode_ltp[subcode]
			for x in lst2:
				flag = False
				if roll in feedback_submitted:
					for row1 in feedback_submitted[roll]:
						#print(row1[1],subcode)
						if row1[1] == subcode:
							#print(ro)
							if int(row1[2]) == int(x):
								flag = True
								break
				#print(flag,subcode,x)
				if flag == False:
					if roll in name_roll:
						ans.append([roll,int(row[0]),int(row[1]),subcode,name_roll[roll][3],name_roll[roll][0],name_roll[roll][1],name_roll[roll][2]])
						break
					else:
						ans.append([roll,int(row[0]),int(row[1]),subcode,'NA_IN_STUDENTINFO','NA_IN_STUDENTINFO','NA_IN_STUDENTINFO','NA_IN_STUDENTINFO'])
						break
					#print(ans)

	'''now we are directly appending our ans to the excel sheet'''
	#print(ans)
	if os.path.isfile('course_feedback_remaining.xlsx'):
		wb = load_workbook(r'course_feedback_remaining.xlsx')
	else:
		wb = Workbook()
	sheet = wb.active
	lst3 = ['rollno','register_sem','schedule_sem','subno','Name','email','aemail','contact']
	sheet.append(lst3)
	for row in ans:
		sheet.append(row)
	wb.save('course_feedback_remaining.xlsx')



 



feedback_not_submitted()
