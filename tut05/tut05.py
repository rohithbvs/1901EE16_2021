import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def generate_marksheet():
    DIRECTORY = "output"
    if not os.path.exists(DIRECTORY): 
        os.makedirs(DIRECTORY)

    '''name_roll dictionary maps roll no to name'''
    name_roll = {}
    f = open('names-roll.csv', 'r')
    with f:
        reader = csv.DictReader(f)

        for row in reader:
            name_roll[row['Roll']] = row['Name']
            #print(row['Name'], row['Country'], row['Profession'])
    #print(name_roll)

    '''sub_dict is a dictionary which takes data from subjects_master and uses subno as key and stored left data as a list'''
    sub_dict = {}
    f = open('subjects_master.csv', 'r')
    with f:
        reader = csv.DictReader(f)

        for row in reader:
            sub_dict[row['subno']] = [row['subname'],row['ltp'],row['crd']]
    #print(sub_dict)

    '''grade is a dictionary of nested list which uses roll no as key and stores roll no related data as list of list'''
    grade = {}
    f = open('grades.csv', 'r')
    with f:
        reader = csv.DictReader(f)

        for row in reader:
            if row['Roll'] in grade:
                grade[row['Roll']].append([row['Sem'],row['SubCode'],row['Credit'],row['Grade'],row['Sub_Type']])
            else:
                grade[row['Roll']]=[]
                grade[row['Roll']].append([row['Sem'],row['SubCode'],row['Credit'],row['Grade'],row['Sub_Type']])
    #print(grade)

    ''' rollno_sem is a dictionary which maps roll no of student to no of semesters he attended'''
    rollno_sem = {}
    for key in grade:
        mset = set()
        for row in grade[key]:
            mset.add(row[0])
            #print(row)
        #print(mset)
        rollno_sem[key] = len(mset)
        mset.clear()
    #print(rollno_sem)
    #rollno_exist = []

    '''the following loop is creating all the .xlsx files with rollno as name of file and no of sheets corresponding to no of sems that student attended'''

    for key in grade:
        wb=Workbook()
        sheet=wb.active
        wb = Workbook()
        #sheet.title = "Overall"
        wb.create_sheet(index=1, title="Overall")
        for i in range(rollno_sem[key]):
            wb.create_sheet(index=i+2, title="Sem"+str(i+1))
        wb.save(f'output\\{key}.xlsx')
   
    '''rollno_credit is a dictionary of list which maps to the roll no of student and total credits taken by student in each sem'''
    '''rollno_spi is a dictionary of list which maps to the roll no of student and spi of that student'''


    rollno_credit = {}
    gradesc = {'AA': 10, 'AB': 9, 'BB': 8, 'BC': 7,
              'CC': 6, 'CD': 5, 'DD': 4, 'F': 0, 'I': 0,'AA*': 10, 'AB*': 9, 'BB*': 8, 'BC*': 7,
              'CC*': 6, 'CD*': 5, 'DD*': 4, 'F*': 0, 'I*': 0, ' BB': 8}
    rollno_spi = {}


    '''the following loops insert the data to corresponding to the sem sheets where they have to be placed'''
    for key in grade:
        wb=load_workbook(r'output\\{}.xlsx'.format(key))
        sheetDelete = wb["Sheet"]
        wb.remove(sheetDelete) 
        rollno_credit[key] = []
        rollno_spi[key] = []
        for i in range(rollno_sem[key]):
            totalcred=0
            spi=0
            name = "Sem"+str(i+1)
            sheet = wb[name]
            header = ["SL No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"]
            sheet.append(header)
            k=1
            for row in grade[key]:
                if int(row[0])==i+1:
                    totalcred=totalcred+int(sub_dict[row[1]][2])
                    spi = spi + int(sub_dict[row[1]][2])*gradesc[row[3]]
                    sheet.append([k,row[1],sub_dict[row[1]][0],sub_dict[row[1]][1],int(sub_dict[row[1]][2]),row[4],row[3]])
                    k=k+1
            rollno_credit[key].append(totalcred)
            if totalcred!=0:
                rollno_spi[key].append(round(spi/totalcred,2))
        wb.save(f'output\\{key}.xlsx')
    #print(rollno_credit)
    #print(rollno_spi)


    '''the following loops calculate the required variables to be placed in overall sheets'''
    rollno_cpi = {}
    rollno_credsofar = {}
    for key in rollno_spi:
        credso = 0
        cpi=0
        rollno_cpi[key] = []
        rollno_credsofar[key] = []
        for i in range(len(rollno_spi[key])):
            cpi = (cpi*credso + rollno_spi[key][i]*rollno_credit[key][i])
            credso = credso + rollno_credit[key][i]
            cpi = round(cpi/credso,2)
            rollno_cpi[key].append(cpi)
            rollno_credsofar[key].append(credso)
    #print(rollno_cpi)
    #print(rollno_credsofar)

    '''the following loop places all the calculated values in overall sheet'''
    for key in grade:
        wb=load_workbook(r'output\\{}.xlsx'.format(key))
        sheet = wb["Overall"]
        sheet["A1"] = "Roll No."
        sheet["B1"] = key
        sheet["A2"] = "Name of Student"
        sheet["B2"] = name_roll[key]
        sheet["A3"] = "Discipline"
        sheet["B3"] = key[4:6]
        sheet["A4"] = "Semester No."
        for i in range(rollno_sem[key]):
            sheet[get_column_letter(i+2)+str(4)] = i+1
        sheet["A5"] = "Semester wise Credit Taken"
        for i in range(len(rollno_credit[key])):
            sheet[get_column_letter(i+2)+str(5)] = rollno_credit[key][i]
        sheet["A6"] = "SPI"
        for i in range(len(rollno_spi[key])):
            sheet[get_column_letter(i+2)+str(6)] = rollno_spi[key][i]
        sheet["A7"] = "Total Credits Taken"
        for i in range(len(rollno_credsofar[key])):
            sheet[get_column_letter(i+2)+str(7)] = rollno_credsofar[key][i]
        sheet["A8"] = "CPI"
        for i in range(len(rollno_cpi[key])):
            sheet[get_column_letter(i+2)+str(8)] = rollno_cpi[key][i]
        wb.save(f'output\\{key}.xlsx')
    return

generate_marksheet()


