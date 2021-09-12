''' my code is working fine but taking a lot of time sir please wait until the code is executed'''
import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook


 
def output_by_subject():
    DIRECTORY = "output_by_subject"
    header = ["rollno","register_sem","subno","sub_type"]
    subno_exist = []

    '''if the directory dosen't exist this is useful for creating'''
    if not os.path.exists(DIRECTORY): 
        os.makedirs(DIRECTORY)
    with open('regtable_old.csv','r') as file:
        studentdata_list=csv.reader(file)
        for data1 in studentdata_list:
            data = []
            data.append(data1[0])
            data.append(data1[1])
            data.append(data1[3])
            data.append(data1[8])
            if (data[2] =="subno"):continue
            if data[2] not in subno_exist:
                subno_exist.append(data[2])
                wb=Workbook()
                sheet=wb.active
                sheet.append(header)
                sheet.append(data)
                wb.save(f'output_by_subject\\{data[2]}.xlsx')
            else:
                wb=load_workbook(r'output_by_subject\\{}.xlsx'.format(data[2]))
                sheet=wb.active
                sheet.append(data)
                wb.save(f'output_by_subject\\{data[2]}.xlsx')
    return
 
def output_individual_roll():
    DIRECTORY = "output_individual_roll"
    header = ["rollno","register_sem","subno","sub_type"]
    rollno_exist = []
    '''if the directory dosen't exist this is useful for creating'''
    if not os.path.exists(DIRECTORY): 
        os.makedirs(DIRECTORY)
    with open('regtable_old.csv','r') as file:
        studentdata_list=csv.reader(file)
        for data1 in studentdata_list:
            data = []
            data.append(data1[0])
            data.append(data1[1])
            data.append(data1[3])
            data.append(data1[8])
            if (data[0] =="rollno"):continue
            write_on_file(data[0],data,rollno_exist,header)
            if data[0] not in rollno_exist:
                rollno_exist.append(data[0])
                wb=Workbook()
                sheet=wb.active
                sheet.append(header)
                sheet.append(data)
                wb.save(f'output_individual_roll\\{data[0]}.xlsx')
            else:
                wb=load_workbook(r'output_individual_roll\\{}.xlsx'.format(data[0]))
                sheet=wb.active
                sheet.append(data)
                wb.save(f'output_individual_roll\\{data[0]}.xlsx')
    return
 
output_by_subject()
output_individual_roll()
