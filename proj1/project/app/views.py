from django.shortcuts import render
from django.http import HttpResponse, response,FileResponse
import pandas as pd
from django.core.files.storage import FileSystemStorage
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import openpyxl
from openpyxl.styles import Font,Fill
from openpyxl.styles import colors
from openpyxl.styles import PatternFill
from django.core.mail import message, send_mail,EmailMessage

# Create your views here.
response_dict = {}
response_dict2 = {}
rollno_res = {}
name_roll = {}
rollno_summ = {}
positive_marks = None
negative_marks = None
total_ques = None
answer = []
absent_list = []
def home(request):
    '''i am taking the files from the user using html form post request and storing them in media folder'''
    if request.method == 'POST':
        global positive_marks
        global negative_marks
        uploaded_file1 = request.FILES['document1']
        fs1 = FileSystemStorage()
        fs1.save(uploaded_file1.name,uploaded_file1)
        uploaded_file2 = request.FILES['document2']
        fs2 = FileSystemStorage()
        fs2.save(uploaded_file2.name,uploaded_file2)
        pmarks = request.POST.get('pmarks')
        if len(pmarks) == 1:
            positive_marks = int(pmarks)
        else:
            positive_marks = float(pmarks)
        nmarks = request.POST.get('nmarks')
        if len(nmarks) == 1:
            negative_marks = int(nmarks)
        else:
            negative_marks = float(nmarks)

        DIRECTORY = "sample_output/marksheet"
        if not os.path.exists(DIRECTORY): 
            os.makedirs(DIRECTORY)
        
        ''' name_roll is dictionary which has roll no as key '''
        global name_roll
        with open('media/'+uploaded_file1.name, 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                name_roll[row['roll']] = row['name']
        

        '''response_dict and response_dict2 is a dictionary taken from response.csv with roll no as key'''
        global response_dict 
        global response_dict2 
        with open('media/'+uploaded_file2.name, 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                response_dict[row['Roll Number']] = [row['Email address'],row['Score'],row['Name'],row['IITP webmail'],row['Phone (10 digit only)']]
                response_dict2[row['Roll Number']] = [row['Timestamp'],row['Email address'],row['Score'],row['Name'],row['IITP webmail'],row['Phone (10 digit only)'],row['Roll Number']]
        #print(response_dict2.keys())
        
        ''' answer is a list which stores right answers of given quiz'''
        ''' rollno_res is a dictionary of rollno as key and a list of students response'''
        global answer
        global rollno_res
        with open('media/'+uploaded_file2.name, 'r') as file:
             reader = csv.reader(file)
             #print(reader)
             for row in reader:
                #print(row)
                rollno_res[row[6]] = row[7:]
                if(row[6]=='ANSWER'):
                    answer = row[7:]


        #print(rollno_res)
        #print(answer)

        '''rollno_summ is a dictionary of rollno as key and list of the students right,wrong,unattempted answers'''
        global rollno_summ
        for key in response_dict2:
            if key is not 'Roll Number':
                wb=Workbook()
                sheet=wb.active
                sheet.title = "quiz"
                wb.save(f'sample_output\\marksheet\\{key}.xlsx')
        total_ques = len(answer)
        rollno_summ = {}
        for key in rollno_res:
            right = 0
            wrong = 0
            unattempted = 0
            for x in range(len(rollno_res[key])):
                if len(rollno_res[key][x])==0:
                    unattempted+=1
                elif rollno_res[key][x]==answer[x]:
                    right+=1
                else:
                    wrong+=1
            rollno_summ[key] = [right,wrong,unattempted]



        '''absent list consists of rollno of students who didn't attempt quiz'''
        global absent_list
        for key in name_roll:
            if key not in response_dict.keys():
                absent_list.append(key)
        
        '''uploading all data in excel files'''
        for key in response_dict2:
            wb=load_workbook(r'sample_output\\marksheet\\{}.xlsx'.format(key))
            ws = wb.worksheets[0]
            img = openpyxl.drawing.image.Image('iitp logo.png')
            img.anchor = 'A1'
            img.height = 90
            img.width = 310 
            ws.add_image(img)
            ws['A6'] = "Name:"
            #print(key)
            #print(name_roll[key])
            if key in name_roll.keys():
                ws['B6'] = name_roll[key]
                ws['B6'].font = Font(bold=True)
                ws['D6'] = "Exam:"
                ws['E6'] = "quiz"
                ws['E6'].font = Font(bold=True)
                ws['A7'] = "Roll Number:"
                ws['B7'] = key
                ws['B7'].font = Font(bold=True)
                heading = ['Right','Wrong','Not Attempt','Max']
                for i,value in enumerate(heading):
                    ws.cell(column=i+2,row=9,value=value)
                    ws[get_column_letter(i+2)+'9'].font = Font(bold=True)
                number = ['No.',rollno_summ[key][0],rollno_summ[key][1],rollno_summ[key][1],total_ques]
                for i,value in enumerate(number):
                    ws.cell(column=i+1,row=10,value=value)
                marking = ['Marking',positive_marks,negative_marks,0]
                for i,value in enumerate(marking):
                    ws.cell(column=i+1,row=11,value=value)
                total = ['Total',positive_marks*rollno_summ[key][0],negative_marks*rollno_summ[key][1],0,str(positive_marks*rollno_summ[key][0]+negative_marks*rollno_summ[key][1])+'/'+str(positive_marks*total_ques)]
                for i,value in enumerate(total):
                    ws.cell(column=i+1,row=12,value=value)
                ws['A15'] = 'Student Ans'
                ws['A15'].font = Font(bold=True)
                ws['B15'] = 'Correct Ans'
                ws['B15'].font = Font(bold=True)
                ws['B10'].font = Font(color='008000')
                ws['B11'].font = Font(color='008000')
                ws['B12'].font = Font(color='008000')
                ws['C10'].font = Font(color='f44336')
                ws['C11'].font = Font(color='f44336')
                ws['C12'].font = Font(color='f44336')
                ws['E12'].font = Font(color=colors.BLUE)
                for i,value in enumerate(answer):
                    ws.cell(column=2,row=16+i,value=value)
                    cell = 'B'+str(16+i)
                    ws[cell].font = Font(color=colors.BLUE)
                for i,value in enumerate(rollno_res[key]):
                    ws.cell(column=1,row=16+i,value=value)
                    cell = 'A'+str(i+16)
                    if rollno_res[key][i] == answer[i]:
                        ws[cell].font = Font(color='008000')
                    else:
                        ws[cell].font = Font(color='f44336')
                wb.save(f'sample_output\\marksheet\\{key}.xlsx')
                
        return render(request,'app/base.html',{'message1':'Excel files are created'})
        
            

    return render(request,'app/base.html')

def send_emails(request):
    '''we are sending mails to students emails using smtp server'''
    for key in response_dict:
        mail = EmailMessage("quiz marks", "PFA", "cs384quizproject@gmail.com", [response_dict[key][0],response_dict[key][3]])
        mail.attach_file(f'sample_output\\marksheet\\{key}.xlsx')
        mail.send(fail_silently=True)
    return render(request,'app/base.html',{'message2':'Emails are being sent'})

def concise_sheet(request):
    '''we are creating concise sheet and adding required data'''
    wb=Workbook()
    sheet=wb.active
    sheet.title = "concise_sheet"
    heading = ['Timestamp','Email address','Goggle_Score','Name','IITP webmail','Phone (10 digit only)','Score_Ater_negative','Roll Number']
    for i in range(len(answer)):
        heading.append('Unnamed: '+str(7+i))
    heading.append('statusAns')
    sheet.append(heading)
    for key in rollno_res:
        if key in name_roll.keys():
            lst = [response_dict2[key][0],response_dict2[key][1],response_dict2[key][2],response_dict2[key][3],response_dict2[key][4],response_dict2[key][5],
                   str(positive_marks*rollno_summ[key][0]+negative_marks*rollno_summ[key][1])+'/'+str(len(answer)*positive_marks),response_dict2[key][6]]
            for x in rollno_res[key]:
                lst.append(x)
            temp = '['+str(rollno_summ[key][0])+','+str(rollno_summ[key][1])+','+str(rollno_summ[key][2])+']'
            lst.append(temp)
            sheet.append(lst)
    wb.save(f'sample_output\\marksheet\\concise_sheet.xlsx')
    

    '''adding absent students in excel sheet'''
    wb=load_workbook(r'sample_output\\marksheet\\concise_sheet.xlsx')
    sheet = wb.active
    for x in absent_list:
        sheet.append(['ABSENT','ABSENT','ABSENT',name_roll[x],'ABSENT','ABSENT','ABSENT',x])
    wb.save(f'sample_output\\marksheet\\concise_sheet.xlsx')
    files = open('sample_output\\marksheet\\concise_sheet.xlsx', 'rb')
    response = FileResponse(files)
    return response

    return render(request,'app/base.html',{'message3':'Concise sheet is generated'})

