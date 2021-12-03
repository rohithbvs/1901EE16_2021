from django.shortcuts import render
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
import csv
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from fpdf import FPDF 
from PIL import Image
import datetime
import shutil

#from create_table_fpdf2 import PDF
uploaded_file1 = None
uploaded_file2 = None
uploaded_file3 = None
stamp_requirement = False
signature_requirement=False
stamp_file = ""
signature_file = ""

def generate_marksheet(startRange,endRange):
    #storing file addresses for each csv file
    #print("hi")
    input_folder="media"
    if not os.path.exists(input_folder):
	    return
    grades_file=os.path.join(input_folder,uploaded_file1.name)
    names_roll_file=os.path.join(input_folder,uploaded_file2.name)
    subjects_master_file=os.path.join(input_folder,uploaded_file3.name)
    DIRECTORY = "transcriptsIITP"
    if not os.path.exists(DIRECTORY):
        os.makedirs(DIRECTORY)
    #print('3.hi')
    
   
    # mapping a grade code to numerical value
    grade_map={"AA":10,"AB":9,"BB":8,"BC":7,"CC":6,"CD":5,"DD":4,"F":0,"I":0,"AA*":10,"AB*":9,"BB*":8,"BC*":7,"CC*":6,"CD*":5,"DD*":4,"F*":0,"I*":0}
    
    # mapping a course code to its [course name ,  L-T-P , credit]
    course_map={}
    with open(subjects_master_file,'r') as file:
        reader=csv.reader(file)
        counter=0
        for row in reader:
            counter+=1
            if(counter<=1):
                continue
            course_map[row[0]]=[row[1],row[2],row[3]]


    #finding the max semester number of a particular rolln number
    max_sem={}
    #add all the courses to a particular roll
    #1901ee12=[[sem_i,subcode_i,credit_i,grade_i,sub_type_i],[sem_j,subcode_j,credit_j,grade_j,sub_type_j]]
    roll_map={}
    with open(grades_file,'r') as file:
        reader=csv.reader(file)
        header=[]
        counter=0
        for row in reader:
            counter+=1
            if counter<=1 or row[0]<startRange or row[0]>endRange:
                continue
            if row[0] in roll_map:
                roll_map[row[0]].append([row[1],row[2],int(row[3]),row[4].strip(),row[5]])
                max_sem[row[0]]=max(max_sem[row[0]],int(row[1]))
            else:
                roll_map[row[0]]=[]
                roll_map[row[0]].append([row[1],row[2],int(row[3]),row[4].strip(),row[5]])
                max_sem[row[0]]=int(row[1])
    
    # making file for each roll number and adding all sheets to them
    with open(names_roll_file,'r') as file:
        reader=csv.reader(file)
        counter=0
        for row in reader:
            counter+=1
            if(counter<=1):
                continue
            if(row[0]<startRange or row[0]>endRange):
                continue
            roll_path = os.path.join(DIRECTORY, row[0]+".xlsx")   
            roll_student=row[0]
            name_student=row[1]
            #appending the new Row to its course file path
            if not os.path.isfile(roll_path):
                wb = Workbook()
                sheet = wb.active
    
                sheet.append(["Roll No.",roll_student,"","","","","","",""])
                sheet.append(["Name of Student",name_student,"","","","","","",""])
                sheet.append(["Discipline",roll_student[4:6],"","","","","","",""])
                sem=["Semester No."]
                for x in range(max_sem[roll_student]):
                    sem.append(x+1)
                sheet.append(sem)
                sheet.title="Overall"
                wb.save(roll_path)
                for x in range(max_sem[roll_student]):
                    sheetname="Sem"+str(x+1)
                    wb.create_sheet(index=x+1,title=sheetname)
                    sheets=wb.sheetnames
                    sem_sheet=wb[sheets[x+1]]
                    sem_sheet.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
                wb.save(roll_path)


    #now adding details of courses for each semester in each roll_number xlxs file
    #also for each student store total credits taken and their corresponding results
    
    for roll in roll_map:
        roll_path = os.path.join(DIRECTORY, str(roll)+".xlsx")   
        wb=load_workbook(roll_path)
        sem_wise_credit_taken=["Semester wise Credit Taken"]
        sem_wise_credit_cleared=["Semester wise Credit Cleared"]
        SPI=["SPI"]
        for i in range(max_sem[roll]):
            sem_wise_credit_taken.append(0)
            sem_wise_credit_cleared.append(0)
            SPI.append(0.0)
        for row in roll_map[roll]:
            semester_no=row[0]
            credits=row[2]
            grade=row[3]
            sheet=wb["Sem"+str(row[0])]
            row_count = sheet.max_row
            newRow=[row_count,row[1],course_map[row[1]][0],course_map[row[1]][1],credits,row[4],grade]
            sheet.append(newRow)
            sem_wise_credit_taken[int(semester_no)]+=int(credits)
            sem_wise_credit_cleared[int(semester_no)]+=grade_map[grade]
            SPI[int(semester_no)]+=float(credits)*grade_map[grade]
        wb.save(roll_path)
        sheet=wb["Overall"]    
        total_credits_taken=["Total Credits Taken"]
        CPI=["CPI"]
        for i in range(max_sem[roll]):
            if(sem_wise_credit_taken[i+1]==0):
                SPI[i+1]=0.00
            else:
                SPI[i+1]/=sem_wise_credit_taken[i+1]*1.0
            if(i==0):
                total_credits_taken.append(sem_wise_credit_taken[1])
                CPI.append(SPI[i+1])
            else:
                total_credits_taken.append(sem_wise_credit_taken[i+1]+total_credits_taken[i])
                if(total_credits_taken[i+1]==0):
                    CPI.append(0.0)
                else:
                    CPI.append((CPI[i]*total_credits_taken[i] +SPI[i+1]*sem_wise_credit_taken[i+1])/total_credits_taken[i+1])
        sheet.append(sem_wise_credit_taken)
        for i in range(max_sem[roll]):
            SPI[i+1]=round(SPI[i+1],2)
        sheet.append(SPI)
        sheet.append(total_credits_taken)
        for i in range(max_sem[roll]):
            CPI[i+1]=round(CPI[i+1],2)
        sheet.append(CPI)
        sheet.append(sem_wise_credit_cleared)
        wb.save(roll_path)
    return



def generate_transcript(stamp_file,signature_file):
    output_folder=r"./transcriptsIITP"
    if not os.path.exists(output_folder):
        return
    
    class PDF(FPDF):
        pass # nothing happens when it is executed. 
    
    for roll_file in os.listdir(output_folder):
        roll=roll_file.split('.')[0]
        roll_file=output_folder+"/"+roll_file
        ## opening the previously created xlsx file using 'load_workbook()' method
        xlsx = openpyxl.load_workbook(roll_file)
        ## getting the sheet to active
        sheet = xlsx.active

        #finding total semester by no of sheets in a xlxs
        total_semester=len(xlsx.sheetnames)-1
        
        #student information
        studentname=sheet.cell(row=2, column=2).value
        yod="20"+roll[0:2]
        programme=""
        if(roll[2:4]=="01"):programme= "Bachelor of Technology"
        elif(roll[2:4]=="21"):programme= "Doctor of Philosphy"
        else:programme= "Master of Technology"
        course=sheet.cell(row=3, column=2).value

        # ########################
        # programme="Master of Technology"
        # ########################
 
        # add a page
         
        pdf=PDF("L","mm","A3") if programme == "Bachelor of Technology" else PDF("L","mm","A4") #page format. A4 is the default value of the format, you don't have to specify it.
        pdf.add_page()

        # Set font: Arial size font
        font=8 if programme == "Bachelor of Technology" else 5.5
        pdf.set_font('Arial', size=font)
        xstart=4
        ystart=5
        xend=430  if programme == "Bachelor of Technology" else 305
        yend=287  if programme == "Bachelor of Technology" else 210

        pdf.image("./image.jpg",xstart,ystart,xend,yend)

        #data scrapping from xlsx
        sem_info=[]             #contains credits taken, credits cleared, spi,cpi
        sem_courses_info=[]     #contains a list of all semeters:each sem has a list of all the courses amd the corresponding grades and their other info
        for sem in range(total_semester+1):
            if sem==0:continue
            sem_info.append([sheet.cell(row=5, column=1+sem).value,sheet.cell(row=6, column=1+sem).value,sheet.cell(row=8, column=1+sem).value,sheet.cell(row=9, column=1+sem).value])
        for sem in range(total_semester+1):
            if(sem==0):continue
            sheet = xlsx.worksheets[sem]
            sem_course=[]
            sem_course.append(["Sub Code","Subject Name","L-T-P","CRD","GRD"])
            for course_count in range(sheet.max_row):
                if(course_count==0):continue
                sem_course.append([sheet.cell(row=1+course_count, column=2).value,sheet.cell(row=1+course_count, column=3).value,sheet.cell(row=1+course_count, column=4).value,sheet.cell(row=1+course_count, column=5).value,sheet.cell(row=1+course_count, column=7).value])
            sem_courses_info.append(sem_course)
        
        
        #setting font for student info
        pdf.set_font_size(font)
        if programme != "Bachelor of Technology":
            #setting roll no
            pdf.set_xy(87.0,33.5)
            pdf.cell(w=15,h=5,txt=roll,border=0)
            
            #setting name of student
            pdf.set_xy(130.0,33.5)
            pdf.cell(w=5,h=5,txt=studentname,border=0)
            
            #setting year of admission
            pdf.set_xy(225.0,33.5)
            pdf.cell(w=10,h=5,txt=yod,border=0)

            #setting programme
            pdf.set_xy(87.5,38)
            pdf.cell(w=20,h=5,txt=programme,border=0)

            #setting programme
            pdf.set_xy(130,38)
            pdf.cell(w=25,h=5,txt=course,border=0)
        else:
            pdf.set_font_size(font+2)
            #setting roll no
            pdf.set_xy(120,45.6)
            pdf.cell(w=15,h=5,txt=roll,border=0)
            
            #setting name of student
            pdf.set_xy(180.0,45.6)
            pdf.cell(w=5,h=5,txt=studentname,border=0)
            
            #setting year of admission
            pdf.set_xy(315.0,45.6)
            pdf.cell(w=10,h=5,txt=yod,border=0)

            #setting programme
            pdf.set_xy(120,51.4)
            pdf.cell(w=20,h=5,txt=programme,border=0)

            #setting course
            pdf.set_xy(180,51.4)
            pdf.cell(w=25,h=5,txt=course,border=0)

            pdf.set_font_size(font)


        #logic for building semester tables
        semester_count=0
        #row_height
        rowh=4 if programme == "Bachelor of Technology" else 3
        sem_space_x=8 if programme == "Bachelor of Technology" else 6
        sem_space_y=10
        x=20  if programme == "Bachelor of Technology" else 17
        y=45
        
        
        if programme == "Bachelor of Technology":
            y=60
        ymax=0
        # Effective page width, or just epw
        epw = pdf.w - 2*pdf.l_margin
        sem_width=epw/3.3                    # total 3 semester in a row : 0.3 added to consider space for gaps between 2 consecutive sem in a row
        for sem in sem_courses_info:
            semester_count+=1
            if(semester_count==4 or semester_count==7):
                y=ymax+10

            if semester_count==1 or semester_count==4 or semester_count==7:
                pdf.set_xy(x,y)
            else:
                pdf.set_xy(x+(sem_width+sem_space_x)*((semester_count-1)%3),y)
            
            pdf.set_font('Arial','U')
            pdf.cell(x,rowh,"Semester "+str(semester_count))
            pdf.set_font('Arial')
            if semester_count>1:
                if semester_count==4 or semester_count==7:
                    pdf.set_xy(x+sem_width*((semester_count-1)%3),y+rowh+2)
                else:
                    pdf.set_xy(x+(sem_width+sem_space_x)*((semester_count-1)%3),y+rowh+2)
            else:
                pdf.set_xy(x,y+rowh+2)
    
            col_width = [0.18*sem_width,0.55*sem_width,0.12*sem_width, 0.075*sem_width, 0.075*sem_width]
            
            #start of sem
            for row in sem:
                col=0
                for datum in row:
                    pdf.cell(col_width[col], rowh, str(datum), border=1,align='C')
                    col+=1
                if semester_count!=1 and semester_count!=4 and semester_count!=7:
                    pdf.set_xy(x+(sem_width+sem_space_x)*((semester_count-1)%3),pdf.get_y()+rowh)
                else:
                    pdf.set_xy(x+(sem_width+sem_space_x)*((semester_count-1)%3),pdf.get_y()+rowh)
                ymax=max(pdf.get_y(),ymax)

            pdf.set_xy(pdf.get_x(),pdf.get_y()+2)
            credits_taken=sem_info[semester_count-1][0]
            credits_cleared=sem_info[semester_count-1][3]
            spi=sem_info[semester_count-1][1]
            cpi=sem_info[semester_count-1][2]
            sem_text="Credits Taken: "+ str(credits_taken)+"     Credits Cleared:  "+ str(credits_cleared)+"      SPI: "+str(spi)+"     CPI: "+str(cpi)
            pdf.cell(sem_width*0.85, rowh,sem_text , border=1,align='C')
            if(semester_count==total_semester or semester_count==3 or semester_count==6 or semester_count==9):
                if semester_count>=1 and semester_count<=3:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)
                elif semester_count>=4 and semester_count<=7:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)
                else:
                    pdf.line(12.5,ymax+9,epw+10-1,ymax+9)



        
        if programme == "Bachelor of Technology" and stamp_file != "":
            pdf.image(stamp_file,145,225,32,32)
        elif (stamp_file != ""):
            pdf.image(stamp_file,100,178,20,20)
        if programme == "Bachelor of Technology" and signature_file != "":
            pdf.image(signature_file,325,235,20,20)
        elif (signature_file != ""):
            pdf.image(signature_file,230,172,12,12)
        pdf.set_xy(15,187)
        if programme == "Bachelor of Technology":
            pdf.set_xy(20,255)
        pdf.set_font("Arial","",10)
        now = datetime.datetime.now()
        
        pdf.cell(w=0,h=2,txt="Date Generated: "+(now.strftime("%d %B %Y, %H:%M")) ,border=0)
        pdf.set_xy(220,183)
        if programme == "Bachelor of Technology":
            pdf.set_xy(320,255)
        pdf.cell(w=0,h=2,txt=" ______________________",border=0)
        pdf.set_xy(220,187)
        if programme == "Bachelor of Technology":
            pdf.set_xy(320,260)
        pdf.cell(w=0,h=1,txt="Asistant Registrar(Academic)",border=0)

        #saving the genearated pdf
        pdf.output(f'transcriptsIITP\\{roll}.pdf','F')
    return



def deleteXLSX(startRange,endRange):
    roll_not_present=[]
    xlsx_file=[]
    output_folder=r"./transcriptsIITP"
    for roll_file in os.listdir(output_folder):
        roll=roll_file.split('.')[0]
        ext=roll_file.split('.')[1]
        if(ext=='xlsx'):
            xlsx_file.append(roll_file)
            
    for file in xlsx_file:
        os.remove(output_folder+"/"+file)
    return


def home(request):
    if request.method == "POST":
        #print(request.FILES.keys())
        global uploaded_file1
        uploaded_file1 = request.FILES['document1']
        fs1 = FileSystemStorage()
        fs1.save(uploaded_file1.name,uploaded_file1)
        global uploaded_file2
        uploaded_file2 = request.FILES['document2']
        fs2 = FileSystemStorage()
        fs2.save(uploaded_file2.name,uploaded_file2)
        global uploaded_file3
        uploaded_file3 = request.FILES['document3']
        fs3 = FileSystemStorage()
        fs3.save(uploaded_file3.name,uploaded_file3)        
        if 'seal' in request.FILES.keys():
            uploaded_file4 = request.FILES['seal']
            fs4 = FileSystemStorage()
            fs4.save(uploaded_file4.name,uploaded_file4)
            global stamp_requirement
            global stamp_file
            stamp_file = os.path.join('media',uploaded_file4.name)
            stamp_requirement = True
        if 'sign' in request.FILES.keys():
            uploaded_file5 = request.FILES['sign']
            fs5 = FileSystemStorage()
            fs5.save(uploaded_file5.name,uploaded_file5)
            global signature_requirement
            global signature_file
            signature_file = os.path.join('media',uploaded_file5.name)
            signature_requirement=True


        return render(request,'app/index.html')
    return render(request,'app/index.html')

def generate_over_range(request):
    if request.method == "POST":
        start_range = request.POST.get('start')
        end_range = request.POST.get('end')
        print(stamp_file,signature_file)
        start_range = start_range.upper()
        end_range = end_range.upper()
        generate_marksheet(start_range,end_range)
        generate_transcript(stamp_file,signature_file)
        deleteXLSX(start_range,end_range)
        return render(request,'app/index.html')

    return render(request,'app/index.html')

def generate_all(request):
    start_range=''
    end_range='ZZZZZZZZZ'
    generate_marksheet(start_range,end_range)
    generate_transcript(stamp_file,signature_file)
    deleteXLSX(start_range,end_range)
    return render(request,'app/index.html')
# Create your views here.
